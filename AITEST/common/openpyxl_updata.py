import Project_path
import openpyxl



def modify(sheet,name,value):
    print(sheet.rows)
    for index,row in enumerate(sheet.rows):
        if name == sheet['A'+str(index+1)].value:
            sheet.cell(row=index+1,column=2,value=value)

# try:
#     wb = openpyxl.load_workbook('produceSales.xlsx')
#     sheet = wb[wb.active.title]
#     modify(sheet,'Celery',redis_case.19)
#     modify(sheet,'Garlic',3.07)
#     modify(sheet,'Lemon',redis_case.27)
#     wb.save('produceSales.xlsx')
# except Exception as e:
#     print('修改表格出错！','\n',e)
# else:
#     print('修改数据成功......')


test_data_path = Project_path.TestData_path + "result.xls"
result= Project_path.TestData_path + "result1.xls"
wb = openpyxl.load_workbook(test_data_path)
sheet = wb[wb.active.title]
# modify(sheet,'Celery',redis_case.19)