# coding: utf-8
from config import base_path,cell_config
import os
import openpyxl
from tools.get_log import GetLog
import datetime
log=GetLog.get_logger()  # 初始化日志对象

class FileTool:
    # 初始化
    def __init__(self,filename,device_type):
        # 组装动态文件路径
        self.old_filename = base_path + os.sep + "data" + os.sep + filename  # 用例文件目录
        nowtimeinfo = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.filename=base_path+os.sep+"result"+os.sep+device_type+"_"+nowtimeinfo+filename # 保存的文件名称
        self.back_excel()  # 备份用例

    # 备份excel文件
    def back_excel(self):
        # 打开文件，获取workbook对象
        oldworkbook = openpyxl.load_workbook(self.old_filename)
        oldworkbook.save(self.filename)
        oldworkbook.close()
        self.workbook=openpyxl.load_workbook(self.filename)

        # 3、获取所有的表单对象
        self.sheets = self.workbook.sheetnames
        # 4、获取当前表的表单对象
        self.sheet = self.workbook[self.sheets[0]]

    # 读取excel文件
    def read_excel(self):
        wholedictinfo = list()
        try:
            log.info("读取用例文件........")
            allsteps = []
            dictinfo = []
            for one_sheet in self.sheets:
                worksheet = self.workbook[one_sheet]
                for i in range(2, worksheet.max_row + 1):
                    caseid = worksheet.cell(row=i, column=cell_config.get("case_id")).value  # 用例编号信息
                    casetitle = worksheet.cell(row=i, column=cell_config.get("case_name")).value  # 用例名称信息
                    if caseid != None:
                        allsteps = []
                        dictinfo = {"case_id": caseid, "case_name": casetitle, "case_catory": one_sheet, "steps": []}
                        wholedictinfo.append(dictinfo)
                    linesinfo = dict()
                    params_value=worksheet.cell(row=i, column=cell_config.get("params")).value
                    if params_value==None:
                        continue
                    linesinfo["step"] = worksheet.cell(row=i, column=cell_config.get("step")).value
                    linesinfo["params"] =params_value
                    linesinfo["x_y"] = [i,cell_config.get("result")]
                    linesinfo["x_y_desc"] = [i, cell_config.get("desc")]
                    allsteps.append(linesinfo)
                    if "steps" in dictinfo:
                        dictinfo["steps"] = allsteps
            log.info("读取用例文件完成........")
            return  wholedictinfo
        except Exception as e:
            log.info("读取用例文件异常,异常信息为:{}".format(e))
            return wholedictinfo


    # 写入exel文件
    # x_y 是一个列表
    def write_excel(self,sheet_name,x_y,msg):
        # x_y 参数的格式为列表，如[2,5]
        self.sheet = self.workbook[sheet_name]
        try:
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)


if __name__ == '__main__':
    f=FileTool("data_case.xlsx")
    d=f.read_excel()
    print(d)
    #f.write_excel("空调本机控制",[3,11],"数据写入成功")
    #f.write_excel("空调本机控制", [4, 11], "数据写入成功22222")
    #f.write_excel("跨机控制", [4, 11], "333333")